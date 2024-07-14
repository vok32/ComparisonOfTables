import os
import platform
import pandas as pd
from tkinter import *
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

def get_next_filename(output_file):
    base_name, ext = os.path.splitext(output_file)
    version = 1
    while os.path.exists(output_file):
        version += 1
        output_file = f"{base_name}_v{version}{ext}"
    return output_file

def remove_empty_rows(sheet):
    rows_to_delete = []
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=False):
        if all(cell.value is None for cell in row):
            rows_to_delete.append(row[0].row)
    for row_num in reversed(rows_to_delete):
        sheet.delete_rows(row_num)

def custom_messagebox(title, message, root):
    window = Toplevel(root)
    window.title(title)

    window.geometry(f"400x125+{root.winfo_x()}+{root.winfo_y()}")

    label = Label(window, text=message)
    label.pack(pady=10)

    accept_button = Button(window, text="Принять", command=window.destroy)
    accept_button.pack(pady=5)

    window.transient()
    window.grab_set()
    window.focus_set()
    window.wait_window()

def compare_excel_tables(file1_path, file2_path, output_path, save_option, key_column, root, position):
    table1 = pd.read_excel(file1_path, engine='openpyxl')
    table2 = pd.read_excel(file2_path, engine='openpyxl')

    columns1 = set(table1.columns)
    columns2 = set(table2.columns)

    if columns1 != columns2:
        custom_messagebox("Ошибка", "Таблицы имеют разные столбцы.\nБудут использованы только общие столбцы.\n\nОтсылка на кнопку отклонить в 1C.", root)

    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Результаты сравнения"

    for col_index, col_name in enumerate(table2.columns, start=1):
        new_sheet.cell(row=1, column=col_index).value = col_name

    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_light_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_light_orange = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")

    table1_dict = table1.set_index(key_column).T.to_dict()

    for index, row in table2.iterrows():
        key_value = row[key_column]
        if key_value in table1_dict:
            table1_row = table1_dict[key_value]
            differences = False
            for col_index, col_name in enumerate(table2.columns, start=1):
                if col_name in table1_row and row[col_name] != table1_row[col_name]:
                    differences = True
                    new_sheet.cell(row=index + 2, column=col_index).fill = fill_green
            if differences and (save_option in ["Все строки", "Только измененные строки", "Новые/измененные строки"]):
                for col_index, col_name in enumerate(table2.columns, start=1):
                    new_sheet.cell(row=index + 2, column=col_index).value = row[col_name]
                    if new_sheet.cell(row=index + 2, column=col_index).fill != fill_green:
                        new_sheet.cell(row=index + 2, column=col_index).fill = fill_light_green
        else:
            if save_option in ["Все строки", "Только новые строки", "Новые/измененные строки"]:
                for col_index, col_name in enumerate(table2.columns, start=1):
                    new_sheet.cell(row=index + 2, column=col_index).value = row[col_name]
                    new_sheet.cell(row=index + 2, column=col_index).fill = fill_yellow

    if save_option == "Все строки":
        for index, row in table2.iterrows():
            for col_index, col_name in enumerate(table2.columns, start=1):
                cell = new_sheet.cell(row=index + 2, column=col_index)
                if cell.value is None:
                    cell.value = row[col_name]

    remove_empty_rows(new_sheet)

    unused_columns1 = columns1 - columns2
    unused_columns2 = columns2 - columns1

    max_rows = new_sheet.max_row

    for col_name in unused_columns1:
        if col_name in table2.columns:
            continue
        col_index = table2.columns.get_loc(col_name) + 1
        for row in range(1, max_rows + 1):
            new_sheet.cell(row=row, column=col_index).fill = fill_light_orange

    for col_name in unused_columns2:
        col_index = table2.columns.get_loc(col_name) + 1
        for row in range(1, max_rows + 1):
            new_sheet.cell(row=row, column=col_index).fill = fill_light_orange

    output_path = get_next_filename(output_path)
    new_workbook.save(output_path)

    show_success_window(output_path, root, position)

def show_success_window(output_path, root, position):
    success_window = Toplevel(root)
    success_window.title("Успех")
    
    root.update_idletasks()
    root_position_x = root.winfo_x()
    root_position_y = root.winfo_y()
    success_window.geometry(f"400x150+{root_position_x}+{root_position_y}")
    
    label = Label(success_window, text=f"Результаты сравнения сохранены в файл:\n{output_path}")
    label.pack(pady=10)
    
    open_folder_button = Button(success_window, text="Открыть папку с файлом", command=lambda: open_output_folder(output_path))
    open_folder_button.pack(pady=5)
    
    close_button = Button(success_window, text="Готово", command=success_window.destroy)
    close_button.pack(pady=5)

def open_output_folder(output_path):
    if platform.system() == "Windows":
        os.system(f'explorer /select,"{os.path.abspath(output_path)}"')
    elif platform.system() == "Darwin":
        os.system(f'open -R "{os.path.abspath(output_path)}"')
    else:
        os.system(f'xdg-open "{os.path.dirname(os.path.abspath(output_path))}"')

def select_files(root):
    position = [root.winfo_x(), root.winfo_y()]

    def select_file1():
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if filename:
            file1_entry.delete(0, END)
            file1_entry.insert(0, filename)

    def select_file2():
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if filename:
            file2_entry.delete(0, END)
            file2_entry.insert(0, filename)

    def select_output_folder():
        foldername = filedialog.askdirectory()
        if foldername:
            output_entry.delete(0, END)
            output_entry.insert(0, os.path.join(foldername, "differences.xlsx"))
        else:
            set_default_output_path()

    def set_default_output_path():
        desktop = os.path.join(os.path.join(os.environ['HOME']), 'Desktop')
        comparison_folder = os.path.join(desktop, "Сравнение выгрузок")
        if not os.path.exists(comparison_folder):
            os.makedirs(comparison_folder)
        output_entry.delete(0, END)
        output_entry.insert(0, os.path.join(comparison_folder, "differences.xlsx"))

    def load_columns(file_path):
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            return list(df.columns)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить столбцы из файла: {file_path}\nОшибка: {str(e)}")
            return []

    def show_columns_selection():
        file1_path = file1_entry.get()
        file2_path = file2_entry.get()
        output_path = output_entry.get()
        save_option = save_option_var.get()

        if not file1_path or not file2_path or not output_path:
            messagebox.showerror("Ошибка", "Не все поля были заполнены.")
            return

        columns_file1 = load_columns(file1_path)
        columns_file2 = load_columns(file2_path)

        if len(columns_file1) <= len(columns_file2):
            columns = columns_file1
            compare_file = "первого файла"
        else:
            columns = columns_file2
            compare_file = "второго файла"

        position = [root.winfo_x(), root.winfo_y()]

        window = Toplevel(root)
        window.title("Выберите столбец для сравнения")
        window.geometry(f"400x150+{position[0]}+{position[1]}")

        label = Label(window, text=f"Выберите столбец для сравнения из {compare_file}:")
        label.pack(pady=10)

        combo = ttk.Combobox(window, values=columns, state="readonly")
        combo.pack(pady=5)
        combo.current(0)

        def on_ok():
            selected_column = combo.get()
            window.destroy()
            compare_excel_tables(file1_path, file2_path, output_path, save_option, selected_column, root, position)

        ok_button = Button(window, text="OK", command=on_ok)
        ok_button.pack(pady=10)

    def show_about(root, position):
        about_window = Toplevel(root)
        about_window.title("О разработчике")
        
        # Центрирование окна "О разработчике" относительно главного окна
        root.update_idletasks()
        root_position_x = root.winfo_x()
        root_position_y = root.winfo_y()
        about_window.geometry(f"500x150+{root_position_x}+{root_position_y}")

        label1 = Label(about_window, text="Программный продукт был разработан для облегчения Вашей работы", padx=10, pady=5)
        label1.pack()

        label2 = Label(about_window, text="Программа создана сотрудником 3 меганаправления, студентом 305 кафедры", padx=10, pady=5)
        label2.pack()

        label3 = Label(about_window, text="и просто хорошим человеком - Матюшенко Романом", padx=10, pady=5)
        label3.pack()

        label4 = Label(about_window, text="Ссылка на GitHub - https://github.com/vok32", padx=10, pady=5)
        label4.pack()

        back_button = Button(about_window, text="Назад", command=about_window.destroy)
        back_button.pack()

    root.title("Сравнение таблиц Excel")

    main_frame = Frame(root)
    main_frame.pack(pady=20)

    Label(main_frame, text="Выберите первый файл Excel:").grid(row=0, column=0, sticky=W, padx=10, pady=5)
    file1_entry = Entry(main_frame, width=50)
    file1_entry.grid(row=0, column=1, padx=10, pady=5)
    Button(main_frame, text="Обзор...", command=select_file1).grid(row=0, column=2, padx=10, pady=5)

    Label(main_frame, text="Выберите второй файл Excel:").grid(row=1, column=0, sticky=W, padx=10, pady=5)
    file2_entry = Entry(main_frame, width=50)
    file2_entry.grid(row=1, column=1, padx=10, pady=5)
    Button(main_frame, text="Обзор...", command=select_file2).grid(row=1, column=2, padx=10, pady=5)

    Label(main_frame, text="Выберите папку для сохранения результата:").grid(row=2, column=0, sticky=W, padx=10, pady=5)
    output_entry = Entry(main_frame, width=50)
    output_entry.grid(row=2, column=1, padx=10, pady=5)
    Button(main_frame, text="Обзор...", command=select_output_folder).grid(row=2, column=2, padx=10, pady=5)

    Label(main_frame, text="Опции сохранения:").grid(row=3, column=0, sticky=W, padx=10, pady=5)
    save_option_var = StringVar(value="Все строки")
    save_options = ["Все строки", "Только измененные строки", "Только новые строки", "Новые/измененные строки"]
    save_option_menu = ttk.Combobox(main_frame, textvariable=save_option_var, values=save_options, state="readonly")
    save_option_menu.grid(row=3, column=1, padx=10, pady=5)
    save_option_menu.current(0)

    Button(main_frame, text="Сравнить таблицы", command=show_columns_selection).grid(row=4, columnspan=3, pady=20)

    Button(main_frame, text="О разработчике", command=show_about).grid(row=5, columnspan=3, pady=10)

    set_default_output_path()

if __name__ == "__main__":
    root = Tk()
    root.geometry("600x400")
    select_files(root)
    root.mainloop()