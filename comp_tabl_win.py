from __future__ import annotations

import os
import re
import subprocess
from collections import defaultdict
from pathlib import Path
from typing import Any, Hashable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from tkinter import (
    BOTH,
    E,
    END,
    LEFT,
    W,
    Button,
    Entry,
    Frame,
    Label,
    Radiobutton,
    BooleanVar,
    StringVar,
    Tk,
    Toplevel,
    filedialog,
    messagebox,
    simpledialog,
    ttk,
)

APP_TITLE = "Сравнение таблиц Excel"
DEFAULT_OUTPUT_NAME = "differences.xlsx"
OUTPUT_SHEET_TITLE = "Результаты сравнения"

SAVE_ALL = "Все строки"
SAVE_NEW = "Новые строки"
SAVE_CHANGED = "Измененные строки"
SAVE_NEW_CHANGED = "Новые/измененные строки"
SAVE_MISSING = "Потеряшки"

SAVE_OPTIONS = (
    SAVE_ALL,
    SAVE_NEW,
    SAVE_CHANGED,
    SAVE_NEW_CHANGED,
    SAVE_MISSING,
)

FILL_YELLOW = PatternFill(start_color="FFEB99", end_color="FFEB99", fill_type="solid")
FILL_LIGHT_GREEN = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
FILL_GREEN = PatternFill(start_color="77DD77", end_color="77DD77", fill_type="solid")
FILL_RED = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")


class ComparisonError(ValueError):
    """Понятная пользователю ошибка входных данных или сравнения."""


def get_next_filename(output_file: str | Path) -> Path:
    """Не перезаписывает существующий файл: name.xlsx -> name_v2.xlsx."""
    output_path = Path(output_file)
    if not output_path.exists():
        return output_path

    version = 2
    while True:
        candidate = output_path.with_name(
            f"{output_path.stem}_v{version}{output_path.suffix}"
        )
        if not candidate.exists():
            return candidate
        version += 1


def read_excel_table(file_path: str | Path) -> pd.DataFrame:
    """Читает первый лист Excel-файла и выдаёт понятные сообщения об ошибках."""
    path = Path(file_path)
    if not path.is_file():
        raise ComparisonError(f"Файл не найден:\n{path}")

    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xls"}:
        raise ComparisonError("Поддерживаются файлы .xlsx, .xlsm и .xls.")

    try:
        if suffix == ".xls":
            return pd.read_excel(path)
        return pd.read_excel(path, engine="openpyxl")
    except ImportError as exc:
        if suffix == ".xls":
            raise ComparisonError(
                "Для чтения старого формата .xls требуется пакет xlrd. "
                "Проще всего пересохранить файл в формате .xlsx."
            ) from exc
        raise ComparisonError(f"Не установлен необходимый модуль: {exc}") from exc
    except Exception as exc:
        raise ComparisonError(f"Не удалось прочитать файл:\n{path}\n\n{exc}") from exc


def is_missing(value: Any) -> bool:
    """Безопасно определяет пустое значение pandas/Excel."""
    try:
        result = pd.isna(value)
    except (TypeError, ValueError):
        return False

    if isinstance(result, bool):
        return result

    try:
        return bool(result)
    except (TypeError, ValueError):
        return False


def values_equal(left: Any, right: Any) -> bool:
    """Считает две пустые ячейки равными."""
    left_missing = is_missing(left)
    right_missing = is_missing(right)
    if left_missing or right_missing:
        return left_missing and right_missing

    try:
        result = left == right
    except Exception:
        return False

    try:
        return bool(result)
    except (TypeError, ValueError):
        return False


def excel_value(value: Any) -> Any:
    """Преобразует NaN/NaT/pd.NA в пустую Excel-ячейку."""
    return None if is_missing(value) else value


def normalize_text(value: Any) -> str:
    """Нормализация текста только для поиска соответствий, не для вывода."""
    if is_missing(value):
        return ""
    text = str(value).replace("\xa0", " ")
    text = " ".join(text.split())
    return text.casefold()


def normalize_contract(value: Any) -> str:
    """Нормализует номер договора для дополнительного сопоставления."""
    if is_missing(value):
        return ""

    # Excel иногда читает целые номера как 123.0.
    if isinstance(value, float) and value.is_integer():
        value = int(value)

    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", "", text)
    return text.casefold()


def soft_value(value: Any) -> Any:
    """Нормализованное значение для оценки похожести строк."""
    if is_missing(value):
        return None

    if isinstance(value, str):
        return normalize_text(value)

    if isinstance(value, float) and value.is_integer():
        return int(value)

    return value


def soft_values_equal(left: Any, right: Any) -> bool:
    return soft_value(left) == soft_value(right)


def validate_match_columns(
    table1: pd.DataFrame,
    table2: pd.DataFrame,
    fio_column: Hashable,
    direction_column: Hashable,
    contract_column: Hashable | None,
) -> None:
    """Проверяет наличие ФИО/направления и допустимость выбранных столбцов."""
    for column, label in (
        (fio_column, "ФИО"),
        (direction_column, "Направление"),
    ):
        if column not in table1.columns or column not in table2.columns:
            raise ComparisonError(
                f"Столбец для поля «{label}» должен присутствовать в обеих таблицах."
            )

    if fio_column == direction_column:
        raise ComparisonError("Для ФИО и направления нужно выбрать разные столбцы.")

    if contract_column is not None:
        if contract_column not in table1.columns or contract_column not in table2.columns:
            raise ComparisonError(
                "Выбранный столбец номера договора должен присутствовать в обеих таблицах."
            )
        if contract_column in {fio_column, direction_column}:
            raise ComparisonError(
                "Столбец номера договора не должен совпадать с ФИО или направлением."
            )

    problems: list[str] = []
    for number, table in ((1, table1), (2, table2)):
        fio_empty = int(table[fio_column].apply(lambda x: normalize_text(x) == "").sum())
        direction_empty = int(
            table[direction_column].apply(lambda x: normalize_text(x) == "").sum()
        )
        if fio_empty:
            problems.append(f"в таблице {number} строк без ФИО: {fio_empty}")
        if direction_empty:
            problems.append(
                f"в таблице {number} строк без направления: {direction_empty}"
            )

    if problems:
        problems_text = "\n• ".join(problems)

        raise ComparisonError(
            "Для надёжного сопоставления ФИО и направление должны быть заполнены:\n• "
            f"{problems_text}"
        )


def make_group_key(
    row: pd.Series,
    fio_column: Hashable,
    direction_column: Hashable,
) -> tuple[str, str]:
    return (
        normalize_text(row[fio_column]),
        normalize_text(row[direction_column]),
    )


def contract_compatible(
    old_row: pd.Series,
    new_row: pd.Series,
    contract_column: Hashable | None,
) -> bool:
    """
    Если в обеих строках есть разные номера договоров, это разные записи.
    Если хотя бы с одной стороны номер пустой, сопоставление разрешено.
    """
    if contract_column is None:
        return True

    old_contract = normalize_contract(old_row[contract_column])
    new_contract = normalize_contract(new_row[contract_column])

    if old_contract and new_contract and old_contract != new_contract:
        return False
    return True


def row_difference_score(
    old_row: pd.Series,
    new_row: pd.Series,
    comparison_columns: list[Hashable],
    fio_column: Hashable,
    direction_column: Hashable,
    contract_column: Hashable | None,
) -> int:
    """Чем меньше число, тем больше строки похожи."""
    score = 0

    for column in comparison_columns:
        if column in {fio_column, direction_column}:
            continue

        if contract_column is not None and column == contract_column:
            old_contract = normalize_contract(old_row[column])
            new_contract = normalize_contract(new_row[column])

            # Пустой номер в новой выгрузке не считаем различием: он мог ещё не подгрузиться.
            if not new_contract and old_contract:
                continue

            if old_contract != new_contract:
                score += 2
            continue

        if not soft_values_equal(old_row[column], new_row[column]):
            score += 1

    return score


def changed_columns_for_rows(
    old_row: pd.Series,
    new_row: pd.Series,
    comparison_columns: list[Hashable],
    fio_column: Hashable,
    direction_column: Hashable,
    contract_column: Hashable | None,
    ignored_change_columns: set[Hashable] | None = None,
) -> set[Hashable]:
    """Возвращает реально изменившиеся общие столбцы."""
    changed: set[Hashable] = set()
    ignored = ignored_change_columns or set()

    for column in comparison_columns:
        # Столбец продолжает участвовать в сопоставлении строк,
        # но его отличие не считается изменением и не подсвечивается.
        if column in ignored:
            continue
        if column in {fio_column, direction_column}:
            if normalize_text(old_row[column]) != normalize_text(new_row[column]):
                changed.add(column)
            continue

        if contract_column is not None and column == contract_column:
            old_contract = normalize_contract(old_row[column])
            new_contract = normalize_contract(new_row[column])

            # Если номер был в старой выгрузке, а в новой временно пустой,
            # это не считаем изменением.
            if old_contract and not new_contract:
                continue

            if old_contract != new_contract:
                changed.add(column)
            continue

        if not values_equal(old_row[column], new_row[column]):
            changed.add(column)

    return changed


def build_row_matches(
    table1: pd.DataFrame,
    table2: pd.DataFrame,
    fio_column: Hashable,
    direction_column: Hashable,
    contract_column: Hashable | None,
    comparison_columns: list[Hashable],
) -> tuple[dict[int, int], set[int], set[int]]:
    """
    Сопоставляет строки старой и новой таблицы.

    Возвращает:
        new_index -> old_index,
        индексы несопоставленных старых строк,
        индексы несопоставленных новых строк.
    """
    old_groups: dict[tuple[str, str], list[int]] = defaultdict(list)
    new_groups: dict[tuple[str, str], list[int]] = defaultdict(list)

    for old_index, row in table1.iterrows():
        old_groups[make_group_key(row, fio_column, direction_column)].append(old_index)

    for new_index, row in table2.iterrows():
        new_groups[make_group_key(row, fio_column, direction_column)].append(new_index)

    matches: dict[int, int] = {}
    unmatched_old: set[int] = set(table1.index)
    unmatched_new: set[int] = set(table2.index)

    all_group_keys = set(old_groups) | set(new_groups)

    for group_key in all_group_keys:
        old_indices = old_groups.get(group_key, [])
        new_indices = new_groups.get(group_key, [])

        if not old_indices or not new_indices:
            continue

        group_unmatched_old = set(old_indices)
        group_unmatched_new = set(new_indices)

        def register_match(new_index: int, old_index: int) -> None:
            matches[new_index] = old_index
            group_unmatched_new.discard(new_index)
            group_unmatched_old.discard(old_index)
            unmatched_new.discard(new_index)
            unmatched_old.discard(old_index)

        # Шаг 1. Точное совпадение по номеру договора, если он заполнен.
        if contract_column is not None:
            old_by_contract: dict[str, list[int]] = defaultdict(list)
            for old_index in old_indices:
                contract = normalize_contract(table1.at[old_index, contract_column])
                if contract:
                    old_by_contract[contract].append(old_index)

            contract_candidates: list[tuple[int, int, int]] = []
            for new_index in new_indices:
                contract = normalize_contract(table2.at[new_index, contract_column])
                if not contract:
                    continue
                for old_index in old_by_contract.get(contract, []):
                    score = row_difference_score(
                        table1.loc[old_index],
                        table2.loc[new_index],
                        comparison_columns,
                        fio_column,
                        direction_column,
                        contract_column,
                    )
                    contract_candidates.append((score, new_index, old_index))

            for _, new_index, old_index in sorted(contract_candidates):
                new_is_unmatched = new_index in group_unmatched_new
                old_is_unmatched = old_index in group_unmatched_old

                if new_is_unmatched and old_is_unmatched:
                    register_match(new_index, old_index)

        # Шаг 2. Ищем полностью совпадающие по остальным полям строки.
        exact_candidates: list[tuple[int, int]] = []
        for new_index in group_unmatched_new:
            new_row = table2.loc[new_index]
            for old_index in group_unmatched_old:
                old_row = table1.loc[old_index]
                if not contract_compatible(old_row, new_row, contract_column):
                    continue

                score = row_difference_score(
                    old_row,
                    new_row,
                    comparison_columns,
                    fio_column,
                    direction_column,
                    contract_column,
                )
                if score == 0:
                    exact_candidates.append((new_index, old_index))

        for new_index, old_index in exact_candidates:
            if new_index in group_unmatched_new and old_index in group_unmatched_old:
                register_match(new_index, old_index)

        # Шаг 3. Для оставшихся строк выбираем наиболее похожие пары.
        # Разные заполненные номера договоров сюда вообще не допускаются.
        similarity_candidates: list[tuple[int, int, int]] = []
        for new_index in group_unmatched_new:
            new_row = table2.loc[new_index]
            for old_index in group_unmatched_old:
                old_row = table1.loc[old_index]
                if not contract_compatible(old_row, new_row, contract_column):
                    continue

                score = row_difference_score(
                    old_row,
                    new_row,
                    comparison_columns,
                    fio_column,
                    direction_column,
                    contract_column,
                )
                similarity_candidates.append((score, new_index, old_index))

        for _, new_index, old_index in sorted(similarity_candidates):
            if new_index in group_unmatched_new and old_index in group_unmatched_old:
                register_match(new_index, old_index)

    return matches, unmatched_old, unmatched_new


def style_row(sheet, row_number: int, column_count: int, fill: PatternFill) -> None:
    for column_number in range(1, column_count + 1):
        sheet.cell(row=row_number, column=column_number).fill = fill


def format_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(text))
        width = min(max(max_length + 2, 10), 50)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def compare_excel_tables(
    file1_path: str | Path,
    file2_path: str | Path,
    output_path: str | Path,
    save_option: str,
    fio_column: Hashable,
    direction_column: Hashable,
    contract_column: Hashable | None = None,
    ignored_change_columns: list[Hashable] | None = None,
) -> tuple[Path, list[Hashable], list[Hashable]]:
    """
    Сравнивает первую (старую) и вторую (новую) таблицы.

    Строки сопоставляются по ФИО + направлению.
    Номер договора используется как дополнительный признак:
      - одинаковый заполненный номер имеет приоритет;
      - разные заполненные номера считаются разными записями;
      - пустой номер в новой выгрузке не мешает сопоставлению.

    Столбцы, которые есть только в старой таблице, автоматически
    добавляются в результат и переносятся для сопоставленных строк.
    """
    if save_option not in SAVE_OPTIONS:
        raise ComparisonError(f"Неизвестный режим сохранения: {save_option}")

    table1 = read_excel_table(file1_path).reset_index(drop=True)
    table2 = read_excel_table(file2_path).reset_index(drop=True)

    if table1.empty and len(table1.columns) == 0:
        raise ComparisonError("В первой таблице не обнаружены столбцы.")
    if table2.empty and len(table2.columns) == 0:
        raise ComparisonError("Во второй таблице не обнаружены столбцы.")

    validate_match_columns(
        table1,
        table2,
        fio_column,
        direction_column,
        contract_column,
    )

    columns1 = list(table1.columns)
    columns2 = list(table2.columns)
    columns1_set = set(columns1)
    columns2_set = set(columns2)

    common_columns = [column for column in columns2 if column in columns1_set]
    only_table1 = [column for column in columns1 if column not in columns2_set]
    only_table2 = [column for column in columns2 if column not in columns1_set]

    ignored_changes = set(ignored_change_columns or [])
    invalid_ignored_columns = [
        column for column in ignored_changes if column not in common_columns
    ]
    if invalid_ignored_columns:
        columns_text = "\n• ".join(map(str, invalid_ignored_columns))

        raise ComparisonError(
            "Исключить из проверки изменений можно только общие столбцы двух таблиц:\n• "
            f"{columns_text}"
        )

    matches, unmatched_old, unmatched_new = build_row_matches(
        table1,
        table2,
        fio_column,
        direction_column,
        contract_column,
        common_columns,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = OUTPUT_SHEET_TITLE

    if save_option == SAVE_MISSING:
        output_columns = columns1
        carry_for_output: list[Hashable] = []
    else:
        # Главное изменение: все столбцы только из старой таблицы
        # автоматически добавляются в обычный результат.
        output_columns = columns2 + only_table1
        carry_for_output = only_table1

    sheet.append([excel_value(column) for column in output_columns])

    def append_row(
        row_values: pd.Series | dict[Hashable, Any],
        status: str,
        changed_columns: set[Hashable] | None = None,
    ) -> None:
        sheet.append([excel_value(row_values.get(column)) for column in output_columns])
        output_row = sheet.max_row

        if status == "new":
            style_row(sheet, output_row, len(output_columns), FILL_YELLOW)
        elif status == "changed":
            style_row(sheet, output_row, len(output_columns), FILL_LIGHT_GREEN)
            for column in changed_columns or set():
                if column in output_columns:
                    column_number = output_columns.index(column) + 1
                    sheet.cell(row=output_row, column=column_number).fill = FILL_GREEN
        elif status == "missing":
            style_row(sheet, output_row, len(output_columns), FILL_RED)

    if save_option == SAVE_MISSING:
        for old_index, row in table1.iterrows():
            if old_index in unmatched_old:
                append_row(row, "missing")
    else:
        for new_index, new_row in table2.iterrows():
            old_index = matches.get(new_index)

            if old_index is None:
                if save_option in {
                    SAVE_ALL,
                    SAVE_NEW,
                    SAVE_NEW_CHANGED,
                }:
                    result_row = new_row.to_dict()
                    result_row.update({column: None for column in carry_for_output})
                    append_row(result_row, "new")
                continue

            old_row = table1.loc[old_index]
            changed_columns = changed_columns_for_rows(
                old_row,
                new_row,
                common_columns,
                fio_column,
                direction_column,
                contract_column,
                ignored_changes,
            )

            result_row = new_row.to_dict()
            result_row.update(
                {
                    column: old_row[column]
                    for column in carry_for_output
                }
            )

            if changed_columns:
                if save_option in {SAVE_ALL, SAVE_CHANGED, SAVE_NEW_CHANGED}:
                    append_row(result_row, "changed", changed_columns)
            elif save_option == SAVE_ALL:
                append_row(result_row, "unchanged")

    format_sheet(sheet)

    final_output_path = get_next_filename(output_path)
    final_output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(final_output_path)

    return final_output_path, only_table1, only_table2


def center_window(
    window: Tk | Toplevel,
    parent: Tk | Toplevel | None = None,
) -> None:
    """Центрирует окно по экрану или относительно родительского окна."""
    window.update_idletasks()
    width = window.winfo_reqwidth()
    height = window.winfo_reqheight()

    if parent is None:
        x = (window.winfo_screenwidth() - width) // 2
        y = (window.winfo_screenheight() - height) // 2
    else:
        parent.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - width) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - height) // 2

    window.geometry(f"+{max(0, x)}+{max(0, y)}")


def center_after_layout(
    window: Tk | Toplevel,
    parent: Tk | Toplevel | None = None,
) -> None:
    """Центрирует окно после расчёта размеров всех элементов интерфейса."""
    window.after_idle(lambda: center_window(window, parent))


def custom_messagebox(title: str, message: str, root: Tk) -> None:
    window = Toplevel(root)
    window.title(title)

    Label(window, text=message, justify="left", wraplength=500).pack(pady=12, padx=12)
    Button(window, text="Принять", command=window.destroy).pack(pady=5)

    window.transient(root)
    center_after_layout(window, root)
    window.grab_set()
    window.focus_set()
    window.wait_window()


def show_success_window(output_path: Path, root: Tk) -> None:
    success_window = Toplevel(root)
    success_window.title("Успех")

    Label(
        success_window,
        text=f"Результаты сравнения сохранены в файл:\n{display_path(output_path)}",
        wraplength=470,
    ).pack(pady=10, padx=10)

    Button(
        success_window,
        text="Открыть папку с файлом",
        command=lambda: open_output_folder(output_path),
    ).pack(pady=5)
    Button(success_window, text="Готово", command=success_window.destroy).pack(pady=5)

    success_window.transient(root)
    center_after_layout(success_window, root)
    success_window.grab_set()
    success_window.focus_set()
    success_window.wait_window()


def open_output_folder(output_path: str | Path) -> None:
    absolute_path = str(Path(output_path).resolve())
    if os.name == "nt":
        subprocess.run(["explorer", "/select,", absolute_path], check=False)
    else:
        subprocess.run(["xdg-open", str(Path(absolute_path).parent)], check=False)


def default_output_path() -> Path:
    desktop = Path.home() / "Desktop"
    base_folder = desktop if desktop.exists() else Path.home()
    comparison_folder = base_folder / "Сравнение таблиц"
    comparison_folder.mkdir(parents=True, exist_ok=True)
    return comparison_folder / DEFAULT_OUTPUT_NAME


def display_path(path: str | Path) -> str:
    """Показывает путь с обычными слешами, не меняя работу с файлами."""
    return str(path).replace("\\", "/")


def sanitize_filename(filename: str) -> str:
    filename = filename.strip()
    if filename.lower().endswith(".xlsx"):
        filename = filename[:-5]
    filename = re.sub(r'[<>:"/\\|?*]', "_", filename).strip(" .")
    if not filename:
        raise ComparisonError("Имя файла не может быть пустым.")
    return f"{filename}.xlsx"


def find_suggested_column_index(
    columns: list[Hashable],
    exact_names: tuple[str, ...],
    contains_names: tuple[str, ...] = (),
) -> int | None:
    normalized = [normalize_text(column) for column in columns]

    for wanted in exact_names:
        wanted_norm = normalize_text(wanted)
        for index, name in enumerate(normalized):
            if name == wanted_norm:
                return index

    for wanted in contains_names:
        wanted_norm = normalize_text(wanted)
        for index, name in enumerate(normalized):
            if wanted_norm in name:
                return index

    return None


def select_files(root: Tk) -> None:
    def select_file(entry: Entry) -> None:
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
        )
        if filename:
            entry.delete(0, END)
            entry.insert(0, display_path(filename))

    def select_output_folder() -> None:
        foldername = filedialog.askdirectory()
        if foldername:
            output_entry.delete(0, END)
            output_entry.insert(
                0, display_path(Path(foldername) / DEFAULT_OUTPUT_NAME)
            )

    def load_columns(file_path: str) -> list[Hashable] | None:
        try:
            return list(read_excel_table(file_path).columns)
        except ComparisonError as exc:
            messagebox.showerror("Ошибка", str(exc), parent=root)
            return None

    def show_columns_selection() -> None:
        file1_path = file1_entry.get().strip()
        file2_path = file2_entry.get().strip()
        output_path = output_entry.get().strip()
        save_option = save_option_var.get()

        if not file1_path or not file2_path or not output_path:
            messagebox.showerror("Ошибка", "Не все поля были заполнены.", parent=root)
            return

        columns_file1 = load_columns(file1_path)
        columns_file2 = load_columns(file2_path)
        if columns_file1 is None or columns_file2 is None:
            return

        columns2_set = set(columns_file2)
        common_columns = [column for column in columns_file1 if column in columns2_set]
        only_file1_columns = [
            column for column in columns_file1 if column not in columns2_set
        ]

        if len(common_columns) < 2:
            messagebox.showerror(
                "Ошибка",
                "В таблицах должно быть минимум два общих столбца: ФИО и направление.",
                parent=root,
            )
            return

        window = Toplevel(root)
        window.title("Настройка сравнения")
        window.transient(root)

        Label(
            window,
            text=(
                "Строки будут сопоставляться по ФИО + направлению.\n"
                "Номер договора используется как дополнительный признак и может быть пустым."
            ),
            justify="left",
            wraplength=520,
        ).pack(pady=(12, 8), padx=15)

        settings_frame = ttk.LabelFrame(window, text="Столбцы сопоставления", padding=10)
        settings_frame.pack(fill="x", padx=18, pady=5)

        Label(settings_frame, text="ФИО:").grid(row=0, column=0, sticky=W, padx=5, pady=5)
        fio_combo = ttk.Combobox(
            settings_frame,
            width=38,
            values=[str(column) for column in common_columns],
            state="readonly",
        )
        fio_combo.grid(row=0, column=1, padx=5, pady=5)

        Label(settings_frame, text="Направление:").grid(
            row=1, column=0, sticky=W, padx=5, pady=5
        )
        direction_combo = ttk.Combobox(
            settings_frame,
            width=38,
            values=[str(column) for column in common_columns],
            state="readonly",
        )
        direction_combo.grid(row=1, column=1, padx=5, pady=5)

        Label(settings_frame, text="Номер договора:").grid(
            row=2, column=0, sticky=W, padx=5, pady=5
        )
        contract_values = ["— не использовать —"] + [
            str(column) for column in common_columns
        ]
        contract_combo = ttk.Combobox(
            settings_frame,
            width=38,
            values=contract_values,
            state="readonly",
        )
        contract_combo.grid(row=2, column=1, padx=5, pady=5)

        fio_index = find_suggested_column_index(
            common_columns,
            ("ФИО", "Ф.И.О.", "ФИО поступающего"),
            ("фио",),
        )
        direction_index = find_suggested_column_index(
            common_columns,
            ("Конкурсная группа", "Направление", "Направление подготовки", "Специальность"),
            ("конкурсн", "направлен", "специальн"),
        )

        fio_combo.current(fio_index if fio_index is not None else 0)
        direction_combo.current(
            direction_index
            if direction_index is not None
            else (1 if len(common_columns) > 1 else 0)
        )
        contract_combo.current(0)

        Label(
            window,
            text=(
                "При необходимости отметьте столбцы, изменения в которых не нужно "
                "подсвечивать. Значения в результате всё равно будут взяты из новой таблицы."
            ),
            justify="left",
            wraplength=520,
        ).pack(pady=(12, 4), padx=18)

        ignore_columns_frame = ttk.LabelFrame(
            window,
            text="Не учитывать при определении изменений",
            padding=(10, 6),
        )
        ignore_columns_frame.pack(fill="x", pady=4, padx=18)

        ignored_column_vars: list[tuple[Hashable, BooleanVar]] = []
        for index, column in enumerate(common_columns):
            ignored_var = BooleanVar(value=False)
            ttk.Checkbutton(
                ignore_columns_frame,
                text=str(column),
                variable=ignored_var,
            ).grid(
                row=index // 2,
                column=index % 2,
                sticky=W,
                padx=(0, 18),
                pady=2,
            )
            ignored_column_vars.append((column, ignored_var))

        if only_file1_columns:
            columns_text = ", ".join(map(str, only_file1_columns))

            Label(
                window,
                text=(
                    "Столбцы только из старой таблицы будут автоматически добавлены "
                    "в результат и перенесены для найденных строк:\n"
                    f"{columns_text}"
                ),
                justify="left",
                wraplength=500,
            ).pack(pady=(12, 4), padx=18)

        def start_comparison() -> None:
            fio_selected_index = fio_combo.current()
            direction_selected_index = direction_combo.current()
            contract_selected_index = contract_combo.current()

            if fio_selected_index < 0 or direction_selected_index < 0:
                messagebox.showerror(
                    "Ошибка",
                    "Выберите столбцы ФИО и направления.",
                    parent=window,
                )
                return

            fio_column = common_columns[fio_selected_index]
            direction_column = common_columns[direction_selected_index]

            contract_column: Hashable | None = None
            if contract_selected_index > 0:
                contract_column = common_columns[contract_selected_index - 1]

            ignored_change_columns = [
                column
                for column, ignored_var in ignored_column_vars
                if ignored_var.get()
            ]

            root.config(cursor="watch")
            window.config(cursor="watch")
            root.update_idletasks()

            try:
                saved_path, only_table1, only_table2 = compare_excel_tables(
                    file1_path,
                    file2_path,
                    output_path,
                    save_option,
                    fio_column,
                    direction_column,
                    contract_column,
                    ignored_change_columns,
                )
            except ComparisonError as exc:
                messagebox.showerror("Ошибка", str(exc), parent=window)
                return
            except Exception as exc:
                messagebox.showerror(
                    "Непредвиденная ошибка",
                    f"Сравнение не выполнено:\n{exc}",
                    parent=window,
                )
                return
            finally:
                root.config(cursor="")
                window.config(cursor="")

            window.destroy()

            show_success_window(saved_path, root)

        Button(window, text="Создать файл", command=start_comparison).pack(pady=12)
        center_after_layout(window, root)

    frame = Frame(root, padx=10, pady=10)
    frame.pack(padx=10, pady=10, fill=BOTH)

    Label(frame, text="Выберите первый файл (исходная таблица):").grid(
        row=0, column=0, sticky=W
    )
    file1_entry = Entry(frame, width=50)
    file1_entry.grid(row=0, column=1, padx=5, pady=5)
    Button(frame, text="Выбрать файл", command=lambda: select_file(file1_entry)).grid(
        row=0, column=2, padx=5, pady=5
    )

    Label(frame, text="Выберите второй файл (новая таблица):").grid(
        row=1, column=0, sticky=W
    )
    file2_entry = Entry(frame, width=50)
    file2_entry.grid(row=1, column=1, padx=5, pady=5)
    Button(frame, text="Выбрать файл", command=lambda: select_file(file2_entry)).grid(
        row=1, column=2, padx=5, pady=5
    )

    Label(frame, text="Выберите папку для сохранения:").grid(
        row=2, column=0, sticky=W
    )
    output_entry = Entry(frame, width=50)
    output_entry.grid(row=2, column=1, padx=5, pady=5)
    try:
        output_entry.insert(0, display_path(default_output_path()))
    except OSError:
        output_entry.insert(
            0, display_path(Path.home() / DEFAULT_OUTPUT_NAME)
        )

    Button(frame, text="Выбрать папку", command=select_output_folder).grid(
        row=2, column=2, padx=5, pady=5
    )

    def update_filename() -> None:
        filename = simpledialog.askstring(
            "Введите имя файла", "Имя файла:", parent=root
        )
        if filename is None:
            return
        try:
            safe_filename = sanitize_filename(filename)
        except ComparisonError as exc:
            messagebox.showerror("Ошибка", str(exc), parent=root)
            return

        folder_path = Path(output_entry.get()).parent
        output_entry.delete(0, END)
        output_entry.insert(0, display_path(folder_path / safe_filename))

    Button(frame, text="Изменить имя файла", command=update_filename).grid(
        row=3, column=2, padx=5, pady=5
    )

    Label(frame, text="Что сохранить в файле:").grid(row=4, column=0, sticky=W)
    save_option_var = StringVar(value=SAVE_ALL)

    for row_number, option in enumerate(SAVE_OPTIONS, start=4):
        if option == SAVE_MISSING:
            label = "Потеряшки"
        else:
            label = option.replace("Новые/измененные", "Новые + изменённые")
        Radiobutton(
            frame,
            text=label,
            variable=save_option_var,
            value=option,
        ).grid(
            row=row_number,
            column=1,
            columnspan=2,
            padx=5,
            pady=5,
            sticky=W,
        )

    Button(root, text="Далее", command=show_columns_selection, width=20).pack(
        pady=10, padx=10
    )

    button_frame = Frame(root)
    button_frame.pack(pady=10)
    Button(
        button_frame,
        text="О разработчике",
        command=lambda: show_developer_info(root),
        width=20,
    ).pack(padx=10, side=LEFT)
    Button(
        button_frame,
        text="О приложении",
        command=lambda: show_app_info(root),
        width=20,
    ).pack(padx=10, side=LEFT)

    Label(frame, text="© 3МН").grid(
        row=4 + len(SAVE_OPTIONS), column=2, sticky=E, pady=10
    )


def show_developer_info(root: Tk) -> None:
    developer_window = Toplevel(root)
    developer_window.title("О разработчике")
    developer_window.transient(root)

    Label(
        developer_window,
        text="Программный продукт был разработан для облегчения Вашей работы",
        padx=10,
        pady=5,
    ).pack()
    Label(
        developer_window,
        text="Программа создана сотрудником 3 меганаправления, студентом 305 кафедры",
        padx=10,
        pady=5,
    ).pack()
    Label(
        developer_window,
        text="в далёком 2024 году.",
        padx=10,
        pady=5,
    ).pack()
    Label(
        developer_window,
        text="GitHub: https://github.com/vok32",
        padx=10,
        pady=5,
    ).pack()
    Button(developer_window, text="Назад", command=developer_window.destroy).pack()
    center_after_layout(developer_window, root)


def show_app_info(root: Tk) -> None:
    info_message = (
        "Логика сопоставления:\n"
        "    ФИО + направление — основной ключ\n"
        "    Номер договора — дополнительный признак\n"
        "    Пустой номер договора в новой таблице допускается\n"
        "    Разные заполненные номера договоров считаются разными записями\n\n"
        "Цвета выделения:\n"
        "    Новые строки: жёлтый\n"
        "    Изменённая строка: светло-зелёный\n"
        "    Изменённая ячейка: зелёный\n"
        "    Строки из первой таблицы, отсутствующие во второй: красный\n\n"
        "Столбцы, которые есть только в старой таблице, автоматически "
        "добавляются в результат для найденных строк."
    )

    info_window = Toplevel(root)
    info_window.title("Информация о приложении")
    info_window.transient(root)

    Label(
        info_window,
        text=info_message,
        justify="left",
        wraplength=550,
    ).pack(pady=10, padx=10)
    Button(info_window, text="Закрыть", command=info_window.destroy).pack(pady=10)
    center_after_layout(info_window, root)


def main() -> None:
    root = Tk()
    root.title(APP_TITLE)

    select_files(root)
    center_after_layout(root)
    root.mainloop()


if __name__ == "__main__":
    main()
