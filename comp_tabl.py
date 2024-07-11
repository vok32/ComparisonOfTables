import os
import pandas as pd
from tkinter import *
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

def get_next_filename(output_file):
    base_name, ext = os.path.splitext(output_file)
    version = 1
    while os.path.exists(output_file):
        version += 1
        output_file = f"{base_name}_v{version}{ext}"
    return output_file

def remove_empty_rows(sheet):
    rows_to_delete = []
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=False):
        if all(cell.value is None for cell in row):
            rows_to_delete.append(row[0].row)
    for row_num in reversed(rows_to_delete):
        sheet.delete_rows(row_num)

def compare_excel_tables(file1_path, file2_path, output_path, save_option, key_column, root):
    # Чтение таблиц из файлов Excel
    table1 = pd.read_excel(file1_path)
    table2 = pd.read_excel(file2_path)

    # Проверка наличия одинаковых столбцов
    if list(table1.columns) != list(table2.columns):
        messagebox.showerror("Ошибка", "Таблицы имеют разные столбцы.")
        return

    # Создаем новый файл для сохранения результатов
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Результаты сравнения"

    # Записываем заголовки столбцов в том же порядке, как в table2
    for col_index, col_name in enumerate(table2.columns, start=1):
        new_sheet.cell(row=1, column=col_index).value = col_name

    # Цвет заливки для выделения различий
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_light_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Создаем словарь для быстрого поиска строк в table1 по значению ключевого столбца
    table1_dict = table1.set_index(key_column).T.to_dict()

    # Обход всех строк второй таблицы и сравнение с соответствующими строками первой таблицы
    for index, row in table2.iterrows():
        key_value = row[key_column]
        if key_value in table1_dict:
            # Найдена строка с таким же ключевым значением, проверяем на изменения
            table1_row = table1_dict[key_value]
            differences = False
            for col_index, col_name in enumerate(table2.columns, start=1):
                if col_name in table1_row and row[col_name] != table1_row[col_name]:
                    differences = True
                    new_sheet.cell(row=index + 2, column=col_index).fill = fill_green
            if differences and (save_option in ["Все строки", "Только измененные строки", "Новые/измененные строки"]):
                for col_index, col_name in enumerate(table2.columns, start=1):
                    new_sheet.cell(row=index + 2, column=col_index).value = row[col_name]
                    if new_sheet.cell(row=index + 2, column=col_index).fill != fill_green:
                        new_sheet.cell(row=index + 2, column=col_index).fill = fill_light_green
        else:
            # Строка с таким ключевым значением отсутствует в table1, выделяем всю строку
            if save_option in ["Все строки", "Только новые строки", "Новые/измененные строки"]:
                for col_index, col_name in enumerate(table2.columns, start=1):
                    new_sheet.cell(row=index + 2, column=col_index).value = row[col_name]
                    new_sheet.cell(row=index + 2, column=col_index).fill = fill_yellow

    if save_option == "Все строки":
        for index, row in table2.iterrows():
            for col_index, col_name in enumerate(table2.columns, start=1):
                cell = new_sheet.cell(row=index + 2, column=col_index)
                if cell.value is None:
                    cell.value = row[col_name]

    # Удаление пустых строк
    remove_empty_rows(new_sheet)

    # Генерация уникального имени файла для сохранения
    output_path = get_next_filename(output_path)

    # Сохранение файла
    new_workbook.save(output_path)
    
    # Отображение окна с сообщением об успешном сохранении
    show_success_window(output_path, root)

def show_success_window(output_path, root):
    success_window = Toplevel(root)
    success_window.title("Успех")
    success_window.geometry("400x150")
    
    label = Label(success_window, text=f"Результаты сравнения сохранены в файл:\n{output_path}")
    label.pack(pady=10)
    
    open_folder_button = Button(success_window, text="Открыть папку с файлом", command=lambda: open_output_folder(output_path))
    open_folder_button.pack(pady=5)
    
    close_button = Button(success_window, text="Готово", command=success_window.destroy)
    close_button.pack(pady=5)

def open_output_folder(output_path):
    os.system(f'explorer /select,"{os.path.abspath(output_path)}"')

def select_files(root):
    def select_file1():
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if filename:
            file1_entry.delete(0, END)
            file1_entry.insert(0, filename)  # Сохраняем полный путь

    def select_file2():
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if filename:
            file2_entry.delete(0, END)
            file2_entry.insert(0, filename)  # Сохраняем полный путь

    def select_output_folder():
        foldername = filedialog.askdirectory()
        if foldername:
            output_entry.delete(0, END)
            # Установим полный путь с расширением .xlsx
            output_entry.insert(0, os.path.join(foldername, "differences.xlsx"))
        else:
            output_entry.delete(0, END)
            desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            comparison_folder = os.path.join(desktop, "Сравнение выгрузок")
            if not os.path.exists(comparison_folder):
                os.makedirs(comparison_folder)
            output_entry.insert(0, os.path.join(comparison_folder, "differences.xlsx"))

    def load_columns(file_path):
        try:
            df = pd.read_excel(file_path)
            return list(df.columns)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить столбцы из файла: {file_path}\nОшибка: {str(e)}")
            return []

    def update_columns_list(event=None):
        file_path = file1_entry.get()
        if not file_path:
            return
        columns = load_columns(file_path)
        key_column_combo['values'] = columns

    def show_columns_selection():
        file1_path = file1_entry.get()
        file2_path = file2_entry.get()
        output_path = output_entry.get()
        save_option = save_option_var.get()

        if not file1_path or not file2_path or not output_path:
            messagebox.showerror("Ошибка", "Не все поля были заполнены.")
            return

        window = Toplevel(root)
        window.title("Выберите столбец для сравнения")
        window.geometry("400x200")

        label = Label(window, text="Выберите столбец для сравнения:")
        label.pack(pady=10)

        key_column_var = StringVar(window)
        columns = load_columns(file1_path)
        key_column_combo = ttk.Combobox(window, width=30, textvariable=key_column_var, values=columns)
        key_column_combo.pack(pady=10)

        def start_comparison():
            key_column = key_column_var.get().strip()
            if not key_column:
                messagebox.showerror("Ошибка", "Выберите столбец для сравнения.")
                return
            window.destroy()
            compare_excel_tables(file1_path, file2_path, output_path, save_option, key_column, root)

        button = Button(window, text="Продолжить", command=start_comparison)
        button.pack(pady=10)

    frame = Frame(root, padx=10, pady=10)
    frame.pack(padx=10, pady=10)

    file1_label = Label(frame, text="Выберите первый файл:")
    file1_label.grid(row=0, column=0, sticky=W)

    file1_entry = Entry(frame, width=50)
    file1_entry.grid(row=0, column=1, padx=5, pady=5)

    file1_button = Button(frame, text="Выбрать файл", command=select_file1)
    file1_button.grid(row=0, column=2, padx=5, pady=5)

    file2_label = Label(frame, text="Выберите второй файл:")
    file2_label.grid(row=1, column=0, sticky=W)

    file2_entry = Entry(frame, width=50)
    file2_entry.grid(row=1, column=1, padx=5, pady=5)

    file2_button = Button(frame, text="Выбрать файл", command=select_file2)
    file2_button.grid(row=1, column=2, padx=5, pady=5)

    output_label = Label(frame, text="Выберите папку для сохранения:")
    output_label.grid(row=2, column=0, sticky=W)

    output_entry = Entry(frame, width=50)
    output_entry.grid(row=2, column=1, padx=5, pady=5)

    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    comparison_folder = os.path.join(desktop, "Сравнение выгрузок")
    if not os.path.exists(comparison_folder):
        os.makedirs(comparison_folder)
    output_entry.insert(0, os.path.join(comparison_folder, "differences.xlsx"))

    output_button = Button(frame, text="Выбрать папку", command=select_output_folder)
    output_button.grid(row=2, column=2, padx=5, pady=5)

    save_label = Label(frame, text="Что сохранить в файле:")
    save_label.grid(row=3, column=0, sticky=W)

    save_option_var = StringVar(value="Все строки")

    save_radio1 = Radiobutton(frame, text="Все строки", variable=save_option_var, value="Все строки")
    save_radio1.grid(row=3, column=1, columnspan=2, padx=5, pady=5, sticky=W)

    save_radio2 = Radiobutton(frame, text="Только новые строки", variable=save_option_var, value="Только новые строки")
    save_radio2.grid(row=4, column=1, columnspan=2, padx=5, pady=5, sticky=W)

    save_radio3 = Radiobutton(frame, text="Только измененные строки", variable=save_option_var, value="Только измененные строки")
    save_radio3.grid(row=5, column=1, columnspan=2, padx=5, pady=5, sticky=W)

    save_radio4 = Radiobutton(frame, text="Новые/измененные строки", variable=save_option_var, value="Новые/измененные строки")
    save_radio4.grid(row=6, column=1, columnspan=2, padx=5, pady=5, sticky=W)

    start_button = Button(root, text="Далее", command=show_columns_selection, width=20)
    start_button.pack(pady=10, padx=10)

    developer_button = Button(root, text="О разработчике", command=show_developer_info, width=20)
    developer_button.pack(pady=10, padx=10)

# О разработчике
def show_developer_info():
    developer_window = Tk()
    developer_window.title("О разработчике")
    developer_label = Label(developer_window, text="Программный продукт был разработан для облегчения Вашей работы", padx=10, pady=5)
    developer_label.pack()
    developer_label = Label(developer_window, text="Разработчик - https://github.com/vok32", padx=10, pady=5)
    developer_label.pack()

    back_button = Button(developer_window, text="Назад", command=developer_window.destroy)
    back_button.pack()

    developer_window.mainloop()

def main():
    root = Tk()
    root.title("Сравнение таблиц Excel")
    root.geometry("700x400")
    select_files(root)
    root.mainloop()

if __name__ == "__main__":
    main()